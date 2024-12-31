# gestion_bodega/views.py

import os
import io
import pandas as pd
from datetime import datetime
from django.conf import settings
from django.urls import reverse
from rest_framework import viewsets, status, filters, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction, IntegrityError
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from rest_framework.views import APIView
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Articulo,  Movimiento, HistorialStock, Categoria, Task, Ubicacion,
    Marca, Modelo, Motivo, Personal, HistorialPrestamo, EstadoArticulo
)
from .serializers import (
    ArticuloSerializer, MovimientoSerializer, HistorialStockSerializer,
    CategoriaSerializer, TaskSerializer, UbicacionSerializer, MarcaSerializer, ModeloSerializer, MotivoSerializer,
    PersonalSerializer, HistorialPrestamoSerializer, EstadoArticuloSerializer, UserSerializer
)

import logging

# Configuración del logger
logger = logging.getLogger(__name__)


class ArticuloViewSet(viewsets.ModelViewSet):
    
    
    
    """
    ViewSet para manejar las operaciones CRUD de Articulos.
    Incluye acciones personalizadas para descargar plantilla e importar artículos desde archivos Excel.
    """
    queryset = Articulo.objects.select_related('categoria', 'marca', 'modelo', 'ubicacion', 'estado').all()
    serializer_class = ArticuloSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'numero_serie', 'codigo_minvu', 'codigo_interno', 'mac']
    ordering_fields = ['nombre', 'stock_minimo', 'stock_actual']
    ordering = ['nombre']

    def perform_create(self, serializer):
        """
        Lógica de creación de un artículo.
        """
        serializer.save()
        logger.info(f"Artículo creado: {serializer.instance.nombre}")

    @action(detail=True, methods=['put', 'patch'], url_path='actualizar-stock')
    def actualizar_stock(self, request, pk=None):
        """
        Actualiza el stock de un artículo existente.
        """
        articulo = self.get_object()
        stock_actual = request.data.get('stock_actual')

        if stock_actual is None:
            logger.warning("Falta el campo 'stock_actual' en la solicitud.")
            return Response({"error": "El campo 'stock_actual' es requerido."}, status=400)

        try:
            stock_actual = int(stock_actual)
            if stock_actual < 0:
                raise ValueError
        except ValueError:
            logger.warning(f"Stock actual inválido proporcionado: {stock_actual}")
            return Response({"error": "El stock actual debe ser un número válido y no negativo."}, status=400)

        with transaction.atomic():
            stock_anterior = articulo.stock_actual
            articulo.stock_actual = stock_actual
            articulo.save()

            HistorialStock.objects.create(
                articulo=articulo,
                tipo_movimiento='Actualización de Stock',
                cantidad=stock_actual,
                stock_anterior=stock_anterior,
                stock_actual=stock_actual,
                usuario=request.user,
                comentario=f"Actualización de stock de {stock_anterior} a {stock_actual}.",
                motivo=None,
                ubicacion=articulo.ubicacion
            )

            articulo_serializer = self.get_serializer(articulo)
            logger.info(f"Stock actualizado para artículo {articulo.nombre}: {stock_anterior} -> {stock_actual}")

            return Response({
                "message": "Stock actualizado correctamente.",
                "articulo": articulo_serializer.data
            }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='plantilla')
    def descargar_plantilla(self, request):
        """
        Descarga una plantilla de Excel para importar artículos con formatos específicos.
        """
        headers = [
            'nombre',
            'stock_actual',
            'stock_minimo',
            'categoria',
            'ubicacion',
            'marca',
            'modelo',
            'estado',
            'numero_serie',
            'mac',
            'codigo_interno',
            'codigo_minvu',
            'descripcion'
        ]

        # Crear un nuevo libro de trabajo y una hoja activa
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla_Articulos"

        # Definir estilos
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Escribir encabezados con estilos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Ajustar ancho de columnas
        column_widths = [20, 15, 15, 15, 15, 15, 15, 15, 20, 17, 15, 15, 25]
        for i, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # Agregar una fila de ejemplo
        ws.append([
            'Ejemplo Artículo',
            10,  # stock_actual
            5,   # stock_minimo
            'Tecnología',
            'Bodega 1',
            'Marca Ejemplo',
            'Modelo Ejemplo',
            'Bueno',
            'SN123456',
            '00:1A:2B:3C:4D:5E',
            'CI78910',
            'CM11213',
            'Descripción del artículo.'
        ])

        # Aplicar estilos a la fila de ejemplo
        for cell in ws[2]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

        # Guardar el libro en un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_articulos.xlsx'
        logger.info("Plantilla de importación descargada correctamente.")
        return response

    @action(detail=False, methods=['post'], url_path='importar')
    def importar_articulos(self, request):
        """
        Importa artículos desde un archivo Excel (.xlsx).
        - Campos requeridos: nombre, stock_actual, stock_minimo, categoria, ubicacion
        - Campos opcionales: marca, modelo, estado, numero_serie, mac, codigo_interno, codigo_minvu, descripcion
        - Los campos opcionales se dejan en blanco si no hay contenido en la celda.
        """
        file = request.FILES.get('file')
        continue_on_errors = request.data.get('continue_on_errors', 'true').lower() == 'true'

        if not file:
            logger.warning("No se ha proporcionado ningún archivo para importar.")
            return Response({"error": "No se ha proporcionado ningún archivo."}, status=status.HTTP_400_BAD_REQUEST)

        # 1) Leer archivo con pandas
        try:
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                df = pd.read_excel(file, engine='openpyxl')
                logger.info(f"Archivo Excel '{file.name}' leído correctamente.")
            else:
                logger.warning("Formato de archivo no soportado para importación.")
                return Response({"error": "Formato no soportado. Usa archivos Excel (.xlsx o .xls)."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error al leer el archivo: {str(e)}")
            return Response({"error": f"Error al leer el archivo: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        # 2) Validar columnas requeridas
        required_fields = ['nombre', 'stock_actual', 'stock_minimo', 'categoria', 'ubicacion']
        missing_columns = [col for col in required_fields if col not in df.columns]
        if missing_columns:
            ejemplo_estructura = {
                "nombre": "Ejemplo Artículo",
                "stock_actual": 10,
                "stock_minimo": 5,
                "categoria": "Tecnología",
                "ubicacion": "Bodega 1"
            }
            logger.warning(f"Faltan columnas obligatorias en el archivo: {', '.join(missing_columns)}.")
            return Response(
                {
                    "error": f"Faltan columnas obligatorias: {', '.join(missing_columns)}.",
                    "ejemplo_formato": ejemplo_estructura
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # 3) Renombrar columnas para que coincidan con los nombres esperados en el código
        column_mapping = {
            'Nombre': 'nombre',
            'nombre': 'nombre',
            'Stock_Actual': 'stock_actual',
            'stock_actual': 'stock_actual',
            'Stock_Minimo': 'stock_minimo',
            'stock_minimo': 'stock_minimo',
            'Categoría': 'categoria',
            'categoría': 'categoria',
            'Ubicación': 'ubicacion',
            'ubicación': 'ubicacion',
            'Marca': 'marca',
            'marca': 'marca',
            'Modelo': 'modelo',
            'modelo': 'modelo',
            'Estado': 'estado',
            'estado': 'estado',
            'N° Serie': 'numero_serie',
            'n° serie': 'numero_serie',
            'MAC': 'mac',
            'mac': 'mac',
            'Cód. Interno': 'codigo_interno',
            'cód. interno': 'codigo_interno',
            'Cód. Minvu': 'codigo_minvu',
            'cód. minvu': 'codigo_minvu',
            'Descripción': 'descripcion',
            'descripción': 'descripcion',
        }

        # Limpiar los nombres de las columnas para evitar espacios en blanco
        df.columns = df.columns.str.strip()

        df.rename(columns=column_mapping, inplace=True)

        # Registrar las columnas después de renombrar para depuración
        logger.debug(f"Columnas después de renombrar: {df.columns.tolist()}")

        # 4) Preprocesar DataFrame para manejar valores faltantes y tipos de datos
        try:
            # Llenar NaN con cadenas vacías para campos opcionales
            df = df.where(pd.notnull(df), '')

            # Convertir columnas numéricas
            numeric_fields = ['stock_actual', 'stock_minimo']
            for field in numeric_fields:
                if field in df.columns:
                    df[field] = pd.to_numeric(df[field], errors='coerce').fillna(0).astype(int)

            # Convertir columnas que deberían ser cadenas a string y limpiar espacios
            string_fields = [
                'nombre', 'categoria', 'ubicacion', 'estado', 'modelo', 'marca',
                'numero_serie', 'codigo_minvu', 'codigo_interno', 'mac', 'descripcion'
            ]
            for field in string_fields:
                if field in df.columns:
                    df[field] = df[field].astype(str).str.strip()
                    # Convertir cadenas vacías a None para campos únicos
                    if field in ['numero_serie', 'mac', 'codigo_interno', 'codigo_minvu']:
                        df[field] = df[field].replace({'': None})

            logger.info("DataFrame preprocesado correctamente.")
            logger.debug(f"Tipos de Datos:\n{df.dtypes}")
        except Exception as e:
            logger.error(f"Error al preprocesar el DataFrame: {str(e)}")
            return Response({"error": f"Error al procesar los datos: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        # 5) Optimizar consultas a la base de datos cargando en caché
        categorias_cache = {c.nombre.lower(): c for c in Categoria.objects.all()}
        ubicaciones_cache = {u.nombre.lower(): u for u in Ubicacion.objects.all()}
        marcas_cache = {m.nombre.lower(): m for m in Marca.objects.all()}
        modelos_cache = {m.nombre.lower(): m for m in Modelo.objects.all()}
        estados_cache = {e.nombre.lower(): e for e in EstadoArticulo.objects.all()}

        errores = []
        creados = 0
        actualizados = 0
        omitidos = 0  # Contador para filas omitidas

        total_filas = len(df)
        logger.info(f"Total de filas a procesar (sin encabezado): {total_filas}")

        # 6) Iterar filas individuales
        for index, row in df.iterrows():
            fila_num = index + 2  # Asumiendo que el encabezado está en la fila 1

            logger.debug(f"Procesando fila {fila_num}: {row.to_dict()}")

            # Verificar que existan los 5 campos requeridos
            nombre_val = row.get('nombre')
            stock_actual_val = row.get('stock_actual')
            stock_minimo_val = row.get('stock_minimo')
            categoria_val = row.get('categoria')
            ubicacion_val = row.get('ubicacion')

            if not nombre_val or stock_actual_val is None or stock_minimo_val is None or not categoria_val or not ubicacion_val:
                omitidos += 1
                logger.warning(f"Fila {fila_num}: Faltan campos requeridos. Se omitirá esta fila.")
                if continue_on_errors:
                    errores.append(f"Fila {fila_num}: Faltan uno o más campos requeridos (nombre, stock_actual, stock_minimo, categoria, ubicacion).")
                    continue
                else:
                    errores.append(f"Fila {fila_num}: Faltan uno o más campos requeridos (nombre, stock_actual, stock_minimo, categoria, ubicacion).")
                    break  # Detener si no se debe continuar

            try:
                nombre_str = str(nombre_val).strip()
                stock_actual_int = int(stock_actual_val)
                stock_minimo_int = int(stock_minimo_val)
                if stock_actual_int < 0:
                    raise ValueError("El stock actual no puede ser negativo.")
                if stock_minimo_int < 0:
                    raise ValueError("El stock mínimo no puede ser negativo.")
            except Exception as e:
                errores.append(f"Fila {fila_num}: Error al convertir datos requeridos - {str(e)}.")
                logger.error(f"Fila {fila_num}: Error al convertir datos requeridos - {str(e)}.")
                if not continue_on_errors:
                    break
                else:
                    omitidos += 1
                    continue

            # 7) Obtener o crear objetos de FK por nombre usando caché
            try:
                # Categoria
                categoria_nombre = str(categoria_val).strip().lower()
                categoria_obj = categorias_cache.get(categoria_nombre)
                if not categoria_obj:
                    categoria_obj, created = Categoria.objects.get_or_create(
                        nombre__iexact=categoria_val,
                        defaults={'nombre': categoria_val}
                    )
                    if created:
                        categorias_cache[categoria_nombre] = categoria_obj
                        logger.info(f"Categoría creada: {categoria_obj.nombre} (ID: {categoria_obj.id})")

                # Ubicacion
                ubicacion_nombre = str(ubicacion_val).strip().lower()
                ubicacion_obj = ubicaciones_cache.get(ubicacion_nombre)
                if not ubicacion_obj:
                    ubicacion_obj, created = Ubicacion.objects.get_or_create(
                        nombre__iexact=ubicacion_val,
                        defaults={'nombre': ubicacion_val}
                    )
                    if created:
                        ubicaciones_cache[ubicacion_nombre] = ubicacion_obj
                        logger.info(f"Ubicación creada: {ubicacion_obj.nombre} (ID: {ubicacion_obj.id})")

                # Marca (opcional)
                marca_val = row.get('marca')
                marca_obj = None
                if marca_val:
                    marca_nombre = str(marca_val).strip().lower()
                    if marca_nombre:
                        marca_obj = marcas_cache.get(marca_nombre)
                        if not marca_obj:
                            marca_obj, created = Marca.objects.get_or_create(
                                nombre__iexact=marca_nombre,
                                defaults={'nombre': marca_val}
                            )
                            if created:
                                marcas_cache[marca_nombre] = marca_obj
                                logger.info(f"Marca creada: {marca_obj.nombre} (ID: {marca_obj.id})")

                # Modelo (opcional)
                modelo_val = row.get('modelo')
                modelo_obj = None
                if modelo_val:
                    modelo_nombre = str(modelo_val).strip().lower()
                    if modelo_nombre:
                        modelo_obj = modelos_cache.get(modelo_nombre)
                        if not modelo_obj:
                            modelo_obj, created = Modelo.objects.get_or_create(
                                nombre__iexact=modelo_nombre,
                                defaults={'nombre': modelo_val}
                            )
                            if created:
                                modelos_cache[modelo_nombre] = modelo_obj
                                logger.info(f"Modelo creado: {modelo_obj.nombre} (ID: {modelo_obj.id})")

                # Estado (opcional)
                estado_val = row.get('estado')
                estado_obj = None
                if estado_val:
                    estado_nombre = str(estado_val).strip().lower()
                    if estado_nombre:
                        estado_obj = estados_cache.get(estado_nombre)
                        if not estado_obj:
                            estado_obj, created = EstadoArticulo.objects.get_or_create(
                                nombre__iexact=estado_nombre,
                                defaults={'nombre': estado_val}
                            )
                            if created:
                                estados_cache[estado_nombre] = estado_obj
                                logger.info(f"Estado creado: {estado_obj.nombre} (ID: {estado_obj.id})")
            except Exception as e:
                errores.append(f"Fila {fila_num}: Error al obtener o crear objetos de FK - {str(e)}.")
                logger.error(f"Fila {fila_num}: Error al obtener o crear objetos de FK - {str(e)}.")
                if not continue_on_errors:
                    break
                else:
                    omitidos += 1
                    continue

            # Campos opcionales
            numero_serie = row.get('numero_serie')
            mac = row.get('mac')
            codigo_interno = row.get('codigo_interno')
            codigo_minvu = row.get('codigo_minvu')
            descripcion = row.get('descripcion')

            # Convertir campos únicos a None si están vacíos
            numero_serie = numero_serie.strip() if isinstance(numero_serie, str) and numero_serie.strip() else None
            mac = mac.strip() if isinstance(mac, str) and mac.strip() else None
            codigo_interno = codigo_interno.strip() if isinstance(codigo_interno, str) and codigo_interno.strip() else None
            codigo_minvu = codigo_minvu.strip() if isinstance(codigo_minvu, str) and codigo_minvu.strip() else None
            descripcion = descripcion.strip() if isinstance(descripcion, str) and descripcion.strip() else None

            # 8) Verificar duplicados en campos únicos: numero_serie, mac, codigo_interno, codigo_minvu
            unique_fields = {
                'numero_serie': numero_serie,
                'mac': mac,
                'codigo_interno': codigo_interno,
                'codigo_minvu': codigo_minvu,
            }

            existing_articulos = Articulo.objects.none()
            for field_name, field_value in unique_fields.items():
                if field_value:
                    existing_articulos |= Articulo.objects.filter(**{field_name: field_value})

            if existing_articulos.exists():
                # **Actualizar** el artículo existente
                articulo_existente = existing_articulos.first()
                articulo_existente.stock_actual = stock_actual_int
                articulo_existente.stock_minimo = stock_minimo_int
                articulo_existente.nombre = nombre_str  # Opcional: actualizar el nombre
                articulo_existente.categoria = categoria_obj
                articulo_existente.ubicacion = ubicacion_obj
                articulo_existente.marca = marca_obj
                articulo_existente.modelo = modelo_obj
                articulo_existente.estado = estado_obj
                articulo_existente.numero_serie = numero_serie if numero_serie else articulo_existente.numero_serie
                articulo_existente.mac = mac if mac else articulo_existente.mac
                articulo_existente.codigo_interno = codigo_interno if codigo_interno else articulo_existente.codigo_interno
                articulo_existente.codigo_minvu = codigo_minvu if codigo_minvu else articulo_existente.codigo_minvu
                articulo_existente.descripcion = descripcion if descripcion else articulo_existente.descripcion

                try:
                    articulo_existente.save()
                    actualizados += 1
                    logger.info(f"Fila {fila_num}: Artículo '{nombre_str}' actualizado con stock actual {stock_actual_int} y stock mínimo {stock_minimo_int}.")
                except Exception as e:
                    errores.append(f"Fila {fila_num}: Error al actualizar el artículo - {str(e)}.")
                    logger.error(f"Fila {fila_num}: Error al actualizar el artículo - {str(e)}.")
                    if not continue_on_errors:
                        break
                    else:
                        omitidos += 1
                        continue
            else:
                # Crear nuevo artículo
                try:
                    Articulo.objects.create(
                        nombre=nombre_str,
                        categoria=categoria_obj,
                        marca=marca_obj,
                        modelo=modelo_obj,
                        ubicacion=ubicacion_obj,
                        estado=estado_obj,
                        numero_serie=numero_serie,
                        mac=mac,
                        codigo_interno=codigo_interno,
                        codigo_minvu=codigo_minvu,
                        descripcion=descripcion,
                        stock_minimo=stock_minimo_int,
                        stock_actual=stock_actual_int
                    )
                    creados += 1
                    logger.info(f"Fila {fila_num}: Artículo '{nombre_str}' creado con stock actual {stock_actual_int} y stock mínimo {stock_minimo_int}.")
                except IntegrityError as ie:
                    errores.append(f"Fila {fila_num}: Error de integridad al crear el artículo - {str(ie)}.")
                    logger.error(f"Fila {fila_num}: Error de integridad al crear el artículo - {str(ie)}.")
                    if not continue_on_errors:
                        break
                    else:
                        omitidos += 1
                        continue
                except Exception as e:
                    errores.append(f"Fila {fila_num}: Error al crear el artículo - {str(e)}.")
                    logger.error(f"Fila {fila_num}: Error al crear el artículo - {str(e)}.")
                    if not continue_on_errors:
                        break
                    else:
                        omitidos += 1
                        continue

        # 10) Preparar la respuesta
        response_data = {
            "creados": creados,
            "actualizados": actualizados,
            "omitidos": omitidos
        }

        if errores:
            response_data["errores"] = errores
            logger.warning(f"Importación completada con errores. Creados: {creados}, Actualizados: {actualizados}, Omitidos: {omitidos}, Errores: {len(errores)}.")
            if not continue_on_errors:
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Informar al usuario que hubo errores pero la importación continuó
                return Response(response_data, status=status.HTTP_200_OK)

        logger.info(f"Importación realizada con éxito. Creados: {creados}, Actualizados: {actualizados}, Omitidos: {omitidos}.")
        return Response(
            {
                "creados": creados,
                "actualizados": actualizados,
                "omitidos": omitidos,
                "mensaje": "Importación realizada con éxito."
            },
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['put', 'patch'], url_path='actualizar-stock-minimo')
    def actualizar_stock_minimo(self, request, pk=None):
        """
        Actualiza el stock mínimo de un artículo existente.
        """
        articulo = self.get_object()
        stock_minimo = request.data.get('stock_minimo')

        if stock_minimo is None:
            logger.warning("Falta el campo 'stock_minimo' en la solicitud.")
            return Response({"error": "El campo 'stock_minimo' es requerido."}, status=400)

        try:
            stock_minimo = int(stock_minimo)
            if stock_minimo < 0:
                raise ValueError
        except ValueError:
            logger.warning(f"Stock mínimo inválido proporcionado: {stock_minimo}")
            return Response({"error": "El stock mínimo debe ser un número válido y no negativo."}, status=400)

        with transaction.atomic():
            stock_anterior = articulo.stock_minimo
            articulo.stock_minimo = stock_minimo
            articulo.save()

            HistorialStock.objects.create(
                articulo=articulo,
                tipo_movimiento='Actualización de Stock Mínimo',
                cantidad=stock_minimo,
                stock_anterior=stock_anterior,
                stock_actual=articulo.stock_actual,
                usuario=request.user,
                comentario=f"Actualización de stock mínimo de {stock_anterior} a {stock_minimo}.",
                motivo=None,
                ubicacion=articulo.ubicacion
            )

            articulo_serializer = self.get_serializer(articulo)
            logger.info(f"Stock mínimo actualizado para artículo {articulo.nombre}: {stock_anterior} -> {stock_minimo}")

            return Response({
                "message": "Stock mínimo actualizado correctamente.",
                "articulo": articulo_serializer.data
            }, status=status.HTTP_200_OK)



class MovimientoViewSet(viewsets.ModelViewSet):
    queryset = Movimiento.objects.all()
    serializer_class = MovimientoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['tipo_movimiento', 'comentario', 'motivo__nombre', 'personal__nombre']
    ordering_fields = ['fecha', 'tipo_movimiento', 'cantidad']
    ordering = ['-fecha']

    def create(self, request, *args, **kwargs):
        logger.debug(f"Datos recibidos para Movimiento: {request.data}")
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            with transaction.atomic():
                movimiento = serializer.save()
                logger.info(f"Movimiento creado: {movimiento.tipo_movimiento} para artículo {movimiento.articulo.nombre}")
        except ValidationError as e:
            logger.error(f"Error de Validación al crear movimiento: {e.message_dict}")
            return Response({"error": e.message_dict}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error General al crear movimiento: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        headers = self.get_success_headers(serializer.data)
        return Response(MovimientoSerializer(movimiento).data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['post'], url_path='anular')
    def anular_movimiento(self, request, pk=None):
        """
        Anula un movimiento específico, ajustando el stock y eliminando el historial correspondiente.
        """
        movimiento = self.get_object()

        if movimiento.tipo_movimiento not in ['Prestamo', 'Regresado']:
            logger.warning(f"Intento de anular movimiento de tipo inválido: {movimiento.tipo_movimiento}")
            return Response({"error": "Solo se pueden anular movimientos de tipo Prestamo o Regresado."}, status=400)

        with transaction.atomic():
            articulo = movimiento.articulo

            if movimiento.tipo_movimiento == 'Prestamo':
                articulo.stock_actual += movimiento.cantidad
                articulo.stock_prestado -= movimiento.cantidad
                if articulo.stock_prestado <= 0:
                    articulo.prestado = False
                articulo.save()

                # Anular HistorialPrestamo
                historial_prestamo = HistorialPrestamo.objects.filter(
                    articulo=articulo,
                    personal=movimiento.personal,
                    cantidad_restante__gte=movimiento.cantidad
                ).order_by('fecha_prestamo').first()

                if historial_prestamo:
                    historial_prestamo.cantidad_restante -= movimiento.cantidad
                    if historial_prestamo.cantidad_restante < 0:
                        historial_prestamo.cantidad_restante = 0
                    historial_prestamo.save()
                    logger.info(f"Historial de préstamo actualizado para artículo {articulo.nombre}.")
                else:
                    logger.error("No se encontró un historial de préstamo correspondiente.")
                    return Response({"error": "No se encontró un historial de préstamo correspondiente."}, status=400)

            elif movimiento.tipo_movimiento == 'Regresado':
                articulo.stock_actual -= movimiento.cantidad
                articulo.stock_prestado += movimiento.cantidad
                if articulo.stock_prestado > 0:
                    articulo.prestado = True
                articulo.save()

                # Anular movimiento relacionado
                if hasattr(movimiento, 'prestamo_relacionado') and movimiento.prestamo_relacionado:
                    historial_prestamo = movimiento.prestamo_relacionado
                    if isinstance(historial_prestamo, HistorialPrestamo):
                        historial_prestamo.cantidad_restante += movimiento.cantidad
                        historial_prestamo.save()
                        logger.info(f"Historial de préstamo relacionado actualizado para artículo {articulo.nombre}.")

            # Eliminar el movimiento
            movimiento.delete()
            logger.info(f"Movimiento anulado y eliminado: {movimiento.id}")

            return Response({"message": "Movimiento anulado correctamente."}, status=status.HTTP_200_OK)


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Categoría creada: {serializer.instance.nombre}")


class UbicacionViewSet(viewsets.ModelViewSet):
    queryset = Ubicacion.objects.all()
    serializer_class = UbicacionSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Ubicación creada: {serializer.instance.nombre}")


class MarcaViewSet(viewsets.ModelViewSet):
    queryset = Marca.objects.all()
    serializer_class = MarcaSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Marca creada: {serializer.instance.nombre}")


class ModeloViewSet(viewsets.ModelViewSet):
    queryset = Modelo.objects.select_related('marca').all()
    serializer_class = ModeloSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Modelo creado: {serializer.instance.nombre}")


class MotivoViewSet(viewsets.ModelViewSet):
    queryset = Motivo.objects.all()
    serializer_class = MotivoSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Motivo creado: {serializer.instance.nombre}")


class PersonalViewSet(viewsets.ModelViewSet):
    queryset = Personal.objects.all()
    serializer_class = PersonalSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'nombre', 'correo_institucional', 'seccion']
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Personal creado: {serializer.instance.nombre}")


class HistorialStockViewSet(viewsets.ModelViewSet):
    queryset = HistorialStock.objects.select_related('articulo', 'usuario', 'motivo', 'ubicacion').all()
    serializer_class = HistorialStockSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['articulo__nombre', 'tipo_movimiento', 'usuario__username']
    ordering_fields = ['fecha']
    ordering = ['-fecha']
    permission_classes = [IsAuthenticated]


class HistorialPrestamoViewSet(viewsets.ModelViewSet):
    queryset = HistorialPrestamo.objects.all()
    serializer_class = HistorialPrestamoSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        with transaction.atomic():
            articulo = serializer.validated_data['articulo']
            # Verificar si el artículo ya está prestado
            prestamos_activos = HistorialPrestamo.objects.filter(
                articulo=articulo,
                fecha_devolucion__isnull=True
            )
            if prestamos_activos.exists():
                raise ValidationError("Este artículo ya está prestado y no ha sido regresado.")

            # Guardar el préstamo
            serializer.save()
            # **Eliminar la creación de Movimiento aquí si ya se maneja en el modelo**
            # Movimiento.objects.create(...)  <-- Eliminado
            logger.info(f"Préstamo realizado: {serializer.instance.articulo.nombre}")


class EstadoArticuloViewSet(viewsets.ModelViewSet):
    queryset = EstadoArticulo.objects.all()
    serializer_class = EstadoArticuloSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Estado de artículo creado: {serializer.instance.nombre}")


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action == 'create':
            return [permissions.AllowAny()]
        return [IsAuthenticated()]

    def get_serializer_context(self):
        return {'request': self.request}

    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Usuario creado: {serializer.instance.username}")


# Vistas adicionales

class UsuarioDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        logger.info(f"Información del usuario solicitada: {user.username}")
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
        })


class ArticuloStockAPIView(generics.RetrieveAPIView):
    queryset = Articulo.objects.all()
    serializer_class = ArticuloSerializer
    lookup_field = 'pk'
    permission_classes = [IsAuthenticated]


class MovimientoHistoryView(generics.ListAPIView):
    serializer_class = MovimientoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        articulo_id = self.kwargs['pk']
        queryset = Movimiento.objects.filter(articulo__id=articulo_id).order_by('-fecha')
        logger.info(f"Historial de movimientos obtenido para artículo ID {articulo_id}.")
        return queryset


class ArticuloListView(generics.ListAPIView):
    queryset = Articulo.objects.all()
    serializer_class = ArticuloSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre']
    ordering_fields = ['nombre', 'stock_actual', 'stock_minimo']
    ordering = ['nombre']
    permission_classes = [IsAuthenticated]


class CambiarEstadoArticuloAPIView(generics.UpdateAPIView):
    queryset = Articulo.objects.all()
    serializer_class = ArticuloSerializer
    lookup_field = 'pk'
    permission_classes = [IsAuthenticated]

    def update(self, request, *args, **kwargs):
        articulo = self.get_object()
        estado_nuevo_id = request.data.get('estado_nuevo')

        if not estado_nuevo_id:
            logger.warning("Falta el campo 'estado_nuevo' en la solicitud de cambio de estado.")
            return Response(
                {"error": "El campo 'estado_nuevo' es requerido."},
                status=status.HTTP_400_BAD_REQUEST
            )

        estado_nuevo = get_object_or_404(EstadoArticulo, id=estado_nuevo_id)

        with transaction.atomic():
            estado_anterior = articulo.estado
            articulo.estado = estado_nuevo
            articulo.save()

            movimiento = Movimiento.objects.create(
                articulo=articulo,
                tipo_movimiento="Cambio de Estado",
                cantidad=0,
                usuario=request.user,
                ubicacion=articulo.ubicacion,
                comentario=f"Cambio de estado de '{estado_anterior.nombre}' a '{estado_nuevo.nombre}'.",
                motivo=None
            )

            movimiento_serializer = MovimientoSerializer(movimiento)
            articulo_serializer = self.get_serializer(articulo)

            logger.info(f"Estado del artículo '{articulo.nombre}' cambiado de '{estado_anterior.nombre}' a '{estado_nuevo.nombre}'.")

            return Response({
                "message": "Estado del artículo cambiado correctamente.",
                "articulo": articulo_serializer.data,
                "movimiento": movimiento_serializer.data
            }, status=status.HTTP_200_OK)


class TaskViewSet(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = Task.objects.filter(user=user)
        logger.info(f"Tareas obtenidas para usuario: {user.username}")
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        logger.info(f"Tarea creada para usuario: {self.request.user.username}")

